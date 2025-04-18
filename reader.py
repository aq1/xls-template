import io

from openpyxl import load_workbook

from common import File, Sheet, Row, Template
from docx import Document


def _do_read_rows(file: File):
    wb = load_workbook(io.BytesIO(file.content))
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue
        rows.append(
            Row(
                row_number=row_number,
                values={str(headers[i]): str(row[i]) for i in range(len(headers))},
            )
        )
    return Sheet(rows=rows)


def _do_read_template(file: File):
    stream = io.BytesIO(file.content)
    stream.seek(0)
    return Template(
        doc=Document(stream),
    )


def read_rows(file: File):
    try:
        return True, _do_read_rows(file)
    except Exception as e:
        return False, f"Failed to read rows {e}"


def read_template(file: File):
    try:
        return True, _do_read_template(file)
    except Exception as e:
        return False, f"Failed to read template {e}"
