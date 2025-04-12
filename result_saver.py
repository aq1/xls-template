from common import Sheet
from log import error, log
from openpyxl import load_workbook
from settings import get_settings


def save_results_to_xlsx(path: str, sheet: Sheet):
    settings = get_settings()
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if settings.result_column_name not in headers:
        headers.append(settings.result_column_name)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

    column_number = headers.index(settings.result_column_name) + 1
    for row in sheet.rows:
        ws.cell(row=row.row_number, column=column_number, value=row.share_url)

    wb.save(path)
